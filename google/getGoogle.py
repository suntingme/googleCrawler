from selenium import webdriver
from bs4 import BeautifulSoup
import re
from openpyxl import load_workbook  # 导入读取excel文件的模块
from openpyxl import Workbook  # 导入新建excel文件的模块
import time
import random



def parseInfo(str):
    soup = BeautifulSoup(str)
    renamelist=''
    namelist =''
    if "</a>" in str:
        nameall = soup.find_all('a')
        for num in range(0, len(nameall)):
            namelist=namelist+nameall[num].string+','
        info = str.replace('…', '').split('-')
        yearinfo = info[len(info) - 2]
        year = yearinfo.split(',')[1]

        nameinfo=info[len(info) - 3]
        nameinfoss=nameinfo.split(',')
        name=nameinfoss[len(nameinfoss)-1]

        if name.find("</a>") == -1:
            renamelist=namelist+name
        else:
            renamelist=namelist[:-1]

    else:
        info_s = soup.string
        info = info_s.replace('…', '').split('-')
        yearinfo = info[1].split(',')
        year = yearinfo[1]
        renamelist=info[0]

    return renamelist,year



xls_read = load_workbook('/Users/ting/Programming/ting/test.xlsx')  # 打开excel文件名为'pyxl_test.xlsx'
wb = Workbook()  # 新建了一张工作表，并默认创建了一张名叫'Sheet'的sheet，
# wb.active.title = 'Sheet'  # 当前活动中的sheet更名

sheet=wb.get_sheet_by_name("Sheet")

browser = webdriver.Chrome()
url='https://scholar.google.com.hk/scholar?q=left+bundle+branch+pacing&hl=zh-CN&as_sdt=0%2C20&as_vis=1&as_ylo=2017&as_yhi='
browser.get(url)
html_source = browser.page_source
ssource1=html_source.replace('<b>', '')
ssource2=ssource1.replace('</b>', '')
soup = BeautifulSoup(ssource2.encode('utf-8'))

gs_rt=soup.find_all('h3',class_='gs_rt')
gs_a=soup.find_all('div',class_='gs_a')
line=1
for num in range(0,9):
    gs_rt_text=gs_rt[num].a
    print(gs_rt_text.string)
    print(gs_a[num].string)  # W Huang, L Su, S Wu, L Xu, F Xiao, X Zhou… - Canadian Journal of …, 2017 - Elsevier
    renamelist, year = parseInfo(str(gs_a[num]))

    sheet.cell(row=line, column=1).value = "".join(gs_rt_text.string)
    sheet.cell(row=line, column=2).value = renamelist
    sheet.cell(row=line, column=3).value = year
    wb.save('/Users/ting/Programming/ting/test.xlsx')
    line=line+1


# # soup.find_all('a',text=re.compile('次数'))


for num in range(1,20):
    print('sleep...60sed')
    time.sleep((10+random.randint(5, 20)))
    url = 'https://scholar.google.com.hk/scholar?start='+str(num)+'0&q=left+bundle+branch+pacing&hl=zh-CN&as_sdt=0,20&as_ylo=2017&as_vis=1'
    browser.get(url)
    html_source = browser.page_source
    ssource1 = html_source.replace('<b>', '')
    ssource2 = ssource1.replace('</b>', '')
    soup = BeautifulSoup(ssource2.encode('utf-8'))
    gs_rt = soup.find_all('h3', class_='gs_rt')
    gs_a = soup.find_all('div', class_='gs_a')

    for num in range(0, 9):
        try:
            gs_rt_text = gs_rt[num].a
            print(gs_rt_text.string)
            print(
                gs_a[num].string)  # W Huang, L Su, S Wu, L Xu, F Xiao, X Zhou… - Canadian Journal of …, 2017 - Elsevier
            if not gs_rt_text.string is None:
                renamelist, year = parseInfo(str(gs_a[num]))

                sheet.cell(row=line, column=1).value = "".join(gs_rt_text.string)
                sheet.cell(row=line, column=2).value = renamelist
                sheet.cell(row=line, column=3).value = year
                wb.save('/Users/ting/Programming/ting/test.xlsx')
                line = line + 1
        except:
            print("exception occors for......".join(gs_rt_text.string))











