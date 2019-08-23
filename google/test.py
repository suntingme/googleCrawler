from openpyxl import load_workbook  # 导入读取excel文件的模块
from openpyxl import Workbook  # 导入新建excel文件的模块
from bs4 import BeautifulSoup



ss='W Huang, L Su, S Wu, L Xu, F Xiao, X Zhou… - Canadian Journal of …, 2017 - Elsevier'

s1=ss.replace('…','').split('-')

s11=s1[0].split(',')

s12=s1[1].split(',')
print(s12[0])
print(s12[1])
xls_read = load_workbook('/Users/ting/Programming/ting/test.xlsx')  # 打开excel文件名为'pyxl_test.xlsx'
# print(xls_read.sheetnames)  # 查看工作表'pyxl_test.xlsx'中的所有sheet名，以列表形式生成

print(xls_read.active)  # 查看文件pyxl_test的活动中sheet
# xls_read.active.title = 'test'  # 将活动中的sheet名称变更为test
# xls_read_sheet = xls_read.active  # 将活动中的sheet赋值给变量
#
#
# for row in range(1, xls_read_sheet.max_row + 1):
#     for col in range(1, xls_read_sheet.max_column + 1):
#         res = xls_read_sheet.cell(row=row, column=col)
#         if res.value:
#             print(res.value, end=' ')
#             # print(xls_read_sheet.cell(row=row,column=col).value,end=' ')
#     print()
#
# xls_read.save('pyxl_test.xlsx')     # 保存该文件
#
# print('=' * 40)

wb = Workbook()  # 新建了一张工作表，并默认创建了一张名叫'Sheet'的sheet，
# print(wb)  # <openpyxl.workbook.workbook.Workbook object at 0x10515bc50>
# print(wb.get_sheet_names())  # 显示wb工作表中所有的sheet，得到一个列表

# wb.create_sheet('Data', index=1)  # 在wb工作表中新建一个名叫'Data'的sheet，该sheet的序号是1
# print(wb.get_sheet_names())
# del wb['Sheet']     #删除wb工作表中名叫'Sheet'的sheet
# print(wb.get_sheet_names())
# print(wb.active)  # 查看wb工作表中活动中的sheet
# print(wb.active.values)  # 将该wb工作表中活动中的sheet的数据形成一个生成器

wb.active.title = 'test_sheet'  # 当前活动中的sheet更名
# print(wb.sheetnames)

num=1
# numa='A'+num
# numb='B'+num
sheet=wb.get_sheet_by_name("test_sheet")
sheet.cell(row=num, column=1).value = s12[0]
sheet.cell(row=num, column=2).value = s12[0]
# wb.active[numa] =
# wb.active[numb] = s12[1]



# print(wb.active['A'+num].value)
# print(wb.active['B1'].value)
# print(wb.active['C1'].value)
wb.save('/Users/ting/Programming/ting/test.xlsx')


s='<div class="gs_a"><a href="/citations?user=CxAco94AAAAJ&amp;hl=zh-CN&amp;oi=sra">AD Arnold</a>, <a href="/citations?user=Hny3-CQAAAAJ&amp;hl=zh-CN&amp;oi=sra">MJ Shun-Shin</a>, D Keene… - Journal of the …, 2018 - onlinejacc.org</div>'

soup =BeautifulSoup(s)

print(soup.a.string)

sd=soup.find_all('a')

for num in range(0, len(sd)):
    print(sd[num].string)

s11=s.replace('…','').split('-')
s123=s11[len(s11)-3]
s1234=s123.split(',')
s5=s1234[len(s1234)-1]
s12=s11[len(s11)-2]

print(s5)